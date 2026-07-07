from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from legacydb_copilot.services.report_generator import InvestigationReport


_INVALID_SHEET_CHARS = set("[]:*?/\\")


def _safe_sheet_name(name: str) -> str:
    """
    Owner: Mukesh Dabi
    Purpose:
        Internal helper for safe sheet name within excel_generator.py.
    
    Input:
        Function parameters declared in the signature.
    
    Output:
        Return value declared by the type hints or route response model.
    
    How it is called:
        Internal callers in excel_generator.py.
    
    Where it fits in the flow:
        Application orchestration -> service function -> structured result for the next workflow step.
    
    Safety considerations:
        Keep tenant/workspace boundaries and do not introduce unsafe database or secret handling.
    """
    safe = "".join("-" if char in _INVALID_SHEET_CHARS else char for char in str(name)).strip()
    return (safe or "Sheet")[:31]


def _unique_sheet_name(wb: Workbook, name: str) -> str:
    """
    Owner: Mukesh Dabi
    Purpose:
        Internal helper for unique sheet name within excel_generator.py.
    
    Input:
        Function parameters declared in the signature.
    
    Output:
        Return value declared by the type hints or route response model.
    
    How it is called:
        Internal callers in excel_generator.py.
    
    Where it fits in the flow:
        Application orchestration -> service function -> structured result for the next workflow step.
    
    Safety considerations:
        Keep tenant/workspace boundaries and do not introduce unsafe database or secret handling.
    """
    safe = _safe_sheet_name(name)
    if safe not in wb.sheetnames:
        return safe
    for suffix_number in range(2, 100):
        suffix = f" {suffix_number}"
        candidate = f"{safe[:31 - len(suffix)]}{suffix}"
        if candidate not in wb.sheetnames:
            return candidate
    return safe[:27] + " 99"


def _style_sheet(ws) -> None:
    """
    Owner: Mukesh Dabi
    Purpose:
        Internal helper for style sheet within excel_generator.py.
    
    Input:
        Function parameters declared in the signature.
    
    Output:
        Return value declared by the type hints or route response model.
    
    How it is called:
        Internal callers in excel_generator.py.
    
    Where it fits in the flow:
        Application orchestration -> service function -> structured result for the next workflow step.
    
    Safety considerations:
        Keep tenant/workspace boundaries and do not introduce unsafe database or secret handling.
    """
    header_fill = PatternFill("solid", fgColor="E8F4F3")
    thin = Side(style="thin", color="CFDBE2")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="15313F")
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(14, max_len + 2), 60)


def write_xlsx(report: InvestigationReport, output_path: Path) -> None:
    """
    Owner: Mukesh Dabi
    Purpose:
        Handles write xlsx within the Database Support AI application flow.
    
    Input:
        Function parameters declared in the signature.
    
    Output:
        Return value declared by the type hints or route response model.
    
    How it is called:
        Investigation, reporting, verification, or knowledge workflows as needed.
    
    Where it fits in the flow:
        Application orchestration -> service function -> structured result for the next workflow step.
    
    Safety considerations:
        Keep tenant/workspace boundaries and do not introduce unsafe database or secret handling.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    summary = report.executive_summary
    rows = [
        ("Investigation ID", report.cover.investigation_id),
        ("Workspace", report.cover.workspace),
        ("Database", report.cover.database),
        ("Generated On", report.cover.generated_on),
        ("Issue Title", summary.issue_title),
        ("Issue Description", summary.issue_description),
        ("Severity", summary.severity),
        ("Business Impact", summary.business_impact),
        ("Confidence Score", f"{summary.confidence_score}%"),
        ("Estimated Root Cause", summary.estimated_root_cause),
        ("Recommendation Summary", summary.recommendation_summary),
        ("Status", summary.status),
    ]
    ws.append(["Field", "Value"])
    for row in rows:
        ws.append(row)
    _style_sheet(ws)

    for section in report.sections:
        for table in section.tables:
            sheet = wb.create_sheet(_unique_sheet_name(wb, table.title))
            sheet.append(table.columns)
            for row in table.rows:
                sheet.append([row.get(column, "") for column in table.columns])
            _style_sheet(sheet)
        if section.sql_blocks:
            sheet = wb.create_sheet(_unique_sheet_name(wb, section.title))
            sheet.append(["Purpose", "Expected Result", "Risk", "SQL"])
            for block in section.sql_blocks:
                sheet.append([block.purpose, block.expected_result, block.risk, block.sql])
            _style_sheet(sheet)
    wb.save(output_path)
